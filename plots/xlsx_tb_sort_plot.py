import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt

import warnings
warnings.filterwarnings("ignore")

def save_sorted_data_to_excel(df, numeric_columns, output_file):
    '''

    Args:
        df:
        numeric_columns:
        output_file:

    Returns:

    '''
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for col in numeric_columns:
            sorted_df = df.sort_values(by=col, ascending=False)  # 按照数值列从高到低排序
            writer.book.create_sheet(f'Sorted by {col}')
            ws = writer.book[f'Sorted by {col}']
            for r_idx, row in enumerate(dataframe_to_rows(sorted_df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    if r_idx > 0 and isinstance(value, str) and value.startswith("http"):
                        cell.hyperlink = value
        writer._save()


def plot_bar_graphs(df, numeric_columns, string_columns, show_label=True):
    '''

    Args:
        df:
        numeric_columns:
        string_columns:

    Returns:

    '''
    plt.rcParams['font.size'] = 8  # 设置字体大小
    for col in numeric_columns:
        for str_col in string_columns:
            sorted_df = df.sort_values(by=col, ascending=False)  # 按照数值列从高到低排序
            plt.figure(figsize=(16, 12))  # 设置图形大小
            ax = sorted_df.plot.bar(x=str_col, y=col, legend=False)
            plt.title(f'Bar plot of {str_col} by {col}')
            plt.xlabel(str_col)
            plt.ylabel(col)
            plt.tight_layout()  # 自适应调整布局

            # 设置x轴和y轴刻度字体大小
            plt.xticks(fontsize=8)
            plt.yticks(fontsize=8)

            if show_label:
                # 在每个条形上添加数值标注
                for i, v in enumerate(sorted_df[col].values):
                    ax.text(i, v, str(v), ha='center', va='bottom')

            plt.savefig(f'{str_col}_vs_{col}_bar_plot.png', format='png')


if __name__ == '__main__':
    input_file = "/Users/gatilin/youtu-work/mllm.xlsx"  # 请替换为你的Excel文件路径
    output_file = "../utils/output/output1.xlsx"  # 输出文件路径
    df = pd.read_excel(input_file)

    numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
    string_columns = df.select_dtypes(include=['object']).columns.tolist()

    if len(numeric_columns) == 0 or len(string_columns) == 0:
        print("请确保数据包含至少一个数值列和一个字符串列。")
    else:
        save_sorted_data_to_excel(df, numeric_columns, output_file)
        plot_bar_graphs(df, numeric_columns, string_columns, show_label=False)

    print("条形图已保存为PNG文件。")
    print("排序完成，结果已保存到", output_file)
